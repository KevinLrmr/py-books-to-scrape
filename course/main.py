import requests
from bs4 import BeautifulSoup
import openpyxl
import re as regex

def preg_float(string):
    return regex.search('([0-9]+(\\.?)[0-9]*)', string).group(0)

def preg_int(string):
    return regex.search('[0-9]+', string).group(0)

def get_book_infos(url, book_key) :
    
    # product_page_url
    product_page_url = url

    # récupération du HTML
    r = requests.get(product_page_url)
    html = BeautifulSoup(r.content, 'html.parser')
    ws[f'A{book_key}'] = product_page_url

    # titre
    title = html.find('div', {'class':'product_main'}).find("h1").get_text()
    ws[f'C{book_key}'] = title

    # universal_product_code
    ws[f'B{book_key}'] = html.find("th", string="UPC").findNext('td').get_text()

    # Price (excl. tax)
    price_excluded_tax = html.find("th", string="Price (excl. tax)").findNext('td').get_text()
    price_excluded_tax = preg_float(price_excluded_tax)
    ws[f'D{book_key}'] =  float(price_excluded_tax)

    # Price (incl. tax)
    price_including_tax = html.find("th", string="Price (incl. tax)").findNext('td').get_text()
    price_including_tax = preg_float(price_including_tax)
    ws[f'E{book_key}'] =  float(price_including_tax)

    # universal_product_code
    availability = html.find("th", string="Availability").findNext('td').get_text()
    availability = preg_int(availability)
    ws[f'F{book_key}'] = int(availability)

    # image_url
    image_url = html.find('img').attrs['src'].replace('../..', 'https://books.toscrape.com')
    ws[f'J{book_key}'] = image_url

    # Description
    description = html.find('div', id="product_description").findNext('p').get_text()
    ws[f'G{book_key}'] = image_url

    # Category
    category = html.find('li', class_="active").findPrevious('a').get_text()
    ws[f'H{book_key}'] = category

    # rating 
    rating = html.find('p', class_="star-rating").attrs['class'][1]
    match rating:
        case "Zero":
            rating = 0
        case "One":
            rating = 1
        case "Two":
            rating = 2
        case "Three":
            rating = 3
        case "Four":
            rating = 4
        case "Five":
            rating = 5

    ws[f'I{book_key}'] = rating

    print(f'{title} téléchargé avec succès')

workbook = openpyxl.Workbook()
domain_name = "https://books.toscrape.com/"
url_website = f'{domain_name}index.html'

col_index = {
    "A" : "product_page_url",
    "B" : "universal_product_code",
    "C" : "title",
    "D" : "price_including_tax",
    "E" : "price_excluding_tax",
    "F" : "number_available",
    "G" : "product_description",
    "H" : "category",
    "I" : "review_rating",
    "J" : "image_url"
}

# récupération du HTML
website = requests.get(url_website)
html_main = BeautifulSoup(website.content, 'html.parser')

key_category = 1

for category in html_main.find('ul', class_="nav-list").find('li').find_all('li') :

    # Permet de ne capter que quelques catégories pour que ce ne soit pas trop long
    # key_category += 1
    # if key_category >= 5 :
    #     break

    category_name = category.get_text().strip()
    print('')
    print(category_name)

    category_url = domain_name+(category.find('a').attrs['href'])
    ws = workbook.create_sheet(category_name)

    for key, value in col_index.items():
        ws[f'{key}1'] = value

    next_page_exists = True
    book_key = 2

    while(next_page_exists) :

        html_category = BeautifulSoup(requests.get(category_url).content, 'html.parser')

        for article in html_category.find_all('article', class_="product_pod") :
            url_article = article.find('a').attrs['href'].replace('../../..', domain_name+'catalogue')
            get_book_infos(url_article, book_key)
            book_key += 1

        next_page = html_category.find('li', class_="next")
        
        if(next_page) :
            category_url = (category_url+(next_page.find('a').attrs['href'])).replace('index.html', '')
            print('')
            print('Passage à la page suivante...')
        else :
            next_page_exists = False
            print('')
            print('Passage à la catégorie suivante...')

workbook.save('../data/books.xlsx')
print('Passage à la catégorie suivante...')
